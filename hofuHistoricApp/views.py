# Create your views here.
from hofuHistoricApp.models import Labordyn
import operator
from django.http import HttpResponse
from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from openpyxl import Workbook
import pickle
from subprocess import Popen, PIPE, STDOUT
import datetime
import time
from django.conf import settings
import os.path


def index(request):
    return render(request, 'hofuHistoricApp/search.html')

def search(request):
    occupList = [Q()]
    genderList = [Q()]
    dateList = [Q()]
    kwargs={}
    occupNames = ["ocCollier",
                  "ocFiller",
                  "ocGutterman",
                  "ocTeamster",
                  "ocBlacksmith",
                  "ocFounder",
                  "ocWoodcutter",
                  "ocHiredGirl",
                  "ocStoveDealer",
                  "ocMason",
                  "ocTailor",
                  "ocHousework",
                  "ocMoulder",
                  "ocCarpenter",
                  "ocShoemaker",
                  "ocCCC"
                  ]
    
    dateFields = ["tenurefrom",
                  "tenurefr2",
                  "tenurefr3"
                  ]
    
    def occup(occupation):
        if occupation in request.GET and request.GET[occupation]:
            occupList.append(Q(occup1__icontains=request.GET[occupation]))
            occupList.append(Q(occup2__icontains=request.GET[occupation]))
            occupList.append(Q(occup3__icontains=request.GET[occupation]))
            occupList.append(Q(occup4__icontains=request.GET[occupation]))
            occupList.append(Q(occup5__icontains=request.GET[occupation]))
            
    def datesAfter(yearAfter):
        for field in dateFields:
            fieldName = field + '__gte'
            dateList.append(Q(**{fieldName:datetime.date(yearAfter,1,1)}))
        
    def datesBefore(yearBefore):
        for field in dateFields:
            fieldName = field + '__lte'
            dateList.append(Q(**{fieldName:datetime.date(yearBefore,12,31)}))
            
    def datesRange(yearAfter,yearBefore):
        for field in dateFields:
            fieldName = field + '__range'
            dateList.append(Q(**{fieldName:[datetime.date(yearAfter,1,1), datetime.date(yearBefore,12,31)]}))

    if 'sexM' in request.GET and request.GET['sexM']:
        sexM = request.GET['sexM']
        genderList.append(Q(gender=sexM))
    if 'sexF' in request.GET and request.GET['sexF']:
        sexF = request.GET['sexF']
        genderList.append(Q(gender=sexF))
    if 'sexC' in request.GET and request.GET['sexC']:
        sexC = request.GET['sexC']
        genderList.append(Q(gender=sexC))
    if 'africanAmerican' in request.GET and request.GET['africanAmerican']:
        kwargs[ 'race' ] = request.GET['africanAmerican']
    if 'firstname' in request.GET and request.GET['firstname']:
        kwargs[ 'firstname__icontains' ] = request.GET['firstname']
    if 'lastname' in request.GET and request.GET['lastname']:
        kwargs[ 'lastname__icontains' ] = request.GET['lastname']
        
    if 'yearAfter' in request.GET and request.GET['yearAfter'] and 'yearBefore' in request.GET and request.GET['yearBefore']:
        datesRange(int(request.GET['yearAfter']),int(request.GET['yearBefore']))
    elif 'yearAfter' in request.GET and request.GET['yearAfter']:
        datesAfter(int(request.GET['yearAfter']))
        
    elif 'yearBefore' in request.GET and request.GET['yearBefore']:
        datesBefore(int(request.GET['yearBefore']))
        
    for occupation in occupNames:
        occup(occupation)
    
    results_list = Labordyn.objects.filter(reduce(operator.or_, occupList), reduce(operator.or_, genderList), reduce(operator.or_, dateList), **kwargs)
    request.session["query_search"] = pickle.dumps(results_list.query)
    return render(request, 'hofuHistoricApp/results.html', { 'results_list': results_list})

def results(request):
    results_list = Labordyn.objects.order_by('id')
    return render(request, 'hofuHistoricApp/results.html', { 'results_list': results_list})

def export(request):
    results_list = Labordyn.objects.order_by('id')
    results_list.query = pickle.loads(request.session['query_search'])
    wb = Workbook(encoding='utf-8')
    ws = wb.active
    fieldnames = ["ID",
                  "Last Name",
                  "First Name",
                  "Gender",
                  "Race",
                  "Occupation 1",
                  "Occupation 2",
                  "Occupation 3",
                  "Occupation 4",
                  "Occupation 5",
                  "TenureFrom",
                  "TenureTo",
                  "TenureFrom2",
                  "TenureTo2",
                  "TenureFrom3",
                  "TenureTo3",
                  "TenureTo3a"
                  "Notes"]
    for col, field in enumerate(fieldnames,start=1):
        ws.cell(row=1,column=col).value=field

    for row, rowdata in enumerate(results_list.values_list(),start=2):
        for col, val in enumerate(rowdata,start=1):
                if val is not None:
                    ws.cell(row=row,column=col).set_explicit_value(value=str(val),data_type='s')
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=export.xlsx'
    wb.save(response)
    return response

def certificate(request, Labordyn_id, name):
    labordyn = get_object_or_404(Labordyn, pk=Labordyn_id)
    return render(request, 'hofuHistoricApp/certificate.html', {'labordyn': labordyn, 'name': name})

def pdf(request, Labordyn_id, name):
    st = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d-%H-%M-%S')
    file_name = '/tmp/' + st + '.pdf'
    url = ('http://127.0.0.1:8000/ledgers/hofuHistoricApp/certificate/' + Labordyn_id + "/" + name)
    js = settings.STATIC_ROOT + "hofuHistoricApp/js/pdf.js"
    phantomjs = os.path.join(os.path.abspath(os.path.join(settings.BASE_DIR, os.pardir)), "/bin/phantomjs")
    Popen([phantomjs, js, url, file_name], stdout=PIPE, stderr=STDOUT).wait()
    result = open(file_name, 'r')
    response =  HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=certificate.pdf'
    return response
